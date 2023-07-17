import time
from openpyxl import load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()

# 인스타그램 로그인 페이지 열기
driver.get('https://www.instagram.com/accounts/login/')

# 로그인 폼 입력
wait = WebDriverWait(driver, 10)  # 최대 10초까지 대기
username_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#loginForm input[name="username"]')))
username_input.send_keys('인스타 아이디')

password_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#loginForm input[name="password"]')))
password_input.send_keys('비밀번호')

time.sleep(2)

# 로그인 버튼 클릭
login_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="loginForm"]/div/div[3]')))
login_button.click()

time.sleep(3)

# Excel 파일 로드
workbook = load_workbook('파일이름.xlsx') #파일이름
sheet = workbook.active

# 각 셀을 순회하며 계정 검색 및 프로필 비활성화 여부 확인
for row in sheet.iter_rows(min_row=5, values_only=True):  # B5 셀부터 시작
    account = row[1]  # B 열

    time.sleep(2)

    # 검색 페이지 열기
    search_url = f'https://www.instagram.com/{account}/'
    driver.get(search_url)

    # 검색 결과 대기
    time.sleep(2)

    # 프로필 활성화 여부 확인
    is_active = True  # 활성화 여부를 저장할 변수
    is_exist = True
    private_indicator = driver.find_elements(By.CLASS_NAME, '_aa_u')
    exist_indicator = driver.find_elements(By.CSS_SELECTOR, '#mount_0_0_Bw > div > div > div.x9f619.x1n2onr6.x1ja2u2z > div > div > div > div.x78zum5.xdt5ytf.x10cihs4.x1t2pt76.x1n2onr6.x1ja2u2z > div.x9f619.xvbhtw8.x78zum5.x168nmei.x13lgxp2.x5pf9jr.xo71vjh.x1uhb9sk.x1plvlek.xryxfnj.x1c4vz4f.x2lah0s.x1q0g3np.xqjyukv.x1qjc9v5.x1oa3qoh.x1qughib > div.x1gryazu.xh8yej3.x10o80wk.x14k21rp.x17snn68.x6osk4m.x1porb0y > section > main > div > div > span')

    if len(private_indicator) > 0 and private_indicator[0].accessible_name == '비공개 계정입니다':
        is_active = False

    if len(exist_indicator) > 0 and exist_indicator[0].accessible_name == '죄송합니다. 페이지를 사용할 수 없습니다.':
        is_exist = False

    if search_url == f'https://www.instagram.com/None/':
        print('instagram Crowling made by Jordy2435')
        driver.quit()
        exit()
        
    if is_active:
        if is_exist:
            print(f'계정 {account}은(는) 계정이 확인되지 않습니다.')
        else:
            print(f'계정 {account}은(는) 공개 상태입니다.')

    else:
        print(f'계정 {account}은(는) 비공개 상태입니다.')

# 웹 드라이버 종료
driver.quit()